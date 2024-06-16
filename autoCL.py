"""
@file: automated_cover_letter.py
@directory: /Automated-Cover-Letters/

Automate the generation of cover letters based on data extracted from an Excel file. 
Each letter is generated with company-specific details and converted to PDF format
in the same directory level.
"""

# Imports
import os
import sys
import shutil

from pathlib import Path
import docx

from datetime import date
from docx2pdf import convert
from openpyxl import load_workbook
from docx.shared import Pt


class ExcelRow:
    """
    Row of data from the Excel file containing cover letter information.
    """

    def __init__(self, company, position, requisitionID, contact):
        self.company = company
        self.position = position
        self.requisitionID = requisitionID
        self.contact = contact

    def __str__(self):
        return f"{self.company} {self.position} {self.requisitionID} {self.contact}"


def manage_docs(company, position):
    """
    Copying a .docx template file and returning the destination path.

    @params:
        company     : Company name for the cover letter
        position    : Position title for the cover letter

    @returns:
        path_of_destination : Path to the generated cover letter document
    """
    
    path_of_company = company.replace(" ", "")
    path_string = os.path.join(os.getcwd(), path_of_company)
    Path(path_string).mkdir(parents=True, exist_ok=True)
    
    path_of_position = position.replace(" ", "")
    path_of_template = os.path.join(os.getcwd(), 'cl_template.docx')
    path_of_destination = os.path.join(path_string, path_of_position + '.docx')
    
    shutil.copy(path_of_template, path_of_destination)
    
    return path_of_destination


def replace_string(filename, key, replacement):
    """
    Replaces placeholders in the specified Word document with data read from Excel file.

    @params:
        filename        : Path to the Word document to be modified
        key             : Placeholder to be replaced
        replacement     : Actual content to replace the placeholder

    @returns:
        NONE
    """
    
    document = docx.Document(filename)
    
    style = document.styles['Normal']
    font = style.font
    font.name = 'Garamond'
    font.size = Pt(12)

    for paragraph in document.paragraphs:
        if key in paragraph.text:
            print('Replacement Located.')
            text = paragraph.text.replace(key, replacement)
            
            paragraph.text = text
            paragraph.style = document.styles['Normal']
            
    document.save(filename)


def convert_to_pdf(destination):
    """
    Converts the specified Word document to PDF format for industry standard.

    @params:
        destination     : Path to the Word document to be converted

    @returns:
        NONE
    """
    
    convert(destination)


def read_excel():
    """
    Reads data from the Excel file containing relevant company and personal information.

    @returns:
        excel_rows     : List of ExcelRow objects, each representing one row of data from .xlsx file
    """
    
    excel_rows = []
    
    my_file = load_workbook(filename='cl_data.xlsx')
    sh = my_file.active

    for file_rows in range(2, sh.max_row + 1):
        row = []
        
        for file_cols in range(1, sh.max_column + 1):
            cell = sh.cell(row=file_rows, column=file_cols)
            row.append(str(cell.value))
            
        content = ExcelRow(row[0], row[1], row[2], row[3])
        excel_rows.append(content)
        
    return excel_rows


def batch_generate(letters):
    """
    Generates cover letters and converts them to PDF for each row of data read from Excel.

    @params:
        letters     : List of ExcelRow objects representing each cover letter to be generated

    @returns:
        NONE
    """
    
    for letter in letters:
        company = letter.company
        position = letter.position
        requisition_id = letter.requisitionID
        contact = letter.contact

        if requisition_id == 'default':
            requisition_id = ""
        else:
            requisition_id = f" (Req. ID: {requisition_id})"

        if contact == 'default':
            contact = "Hiring Manager"

        placeholders = {
            '#company#': company,
            '#date#': date.today().strftime("%B %d, %Y"),
            '#position#': position,
            '#requisitionID#': requisition_id,
            '#contact#': contact
        }

        destination = manage_docs(company, position)

        for key, value in placeholders.items():
            replace_string(destination, key, value)

        convert_to_pdf(destination)


if __name__ == "__main__":
    os.chdir(os.path.dirname(sys.argv[0]))
    letters = read_excel()
    batch_generate(letters)
